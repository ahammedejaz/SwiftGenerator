"""Excel as a first-class automation input.

Templates are generated *from the message specification*, so an automation tester never
has to invent a tag, a qualifier or an XPath. Every template ships three sheets:

``Scenarios``   the rows you edit — pre-filled with a working example
``Reference``   every field the message supports, with presence, format and an example
``Read me``     what each column means and how to build repeated sequences

Parsing is deliberately forgiving about presentation (case, whitespace, Excel turning a
date into a datetime) and strict about meaning: an unknown tag or XPath is a row-level
error naming the row, not a silent omission.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import is_zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from app.studio.catalogue import message_spec
from app.studio.models import (
    ElementInput,
    FieldInput,
    InputKind,
    IssueSeverity,
    MessageFormat,
    Presence,
    SampleVariant,
    SpecField,
    ValidationIssue,
    ValidationLayer,
)
from app.studio.samples import build_sample

HEADER_FILL = PatternFill("solid", fgColor="0F3D3E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
MANDATORY_FILL = PatternFill("solid", fgColor="FDECEC")
GROUP_FONT = Font(bold=True)

MT_HEADERS = [
    "ScenarioID",
    "MessageType",
    "ProfileID",
    "Sequence",
    "SequenceOccurrence",
    "Tag",
    "Qualifier",
    "Option",
    "Value",
]
MX_HEADERS = [
    "ScenarioID",
    "MessageType",
    "ProfileID",
    "XPath",
    "Occurrence",
    "Value",
]

MT_REQUIRED = {"ScenarioID", "MessageType", "Tag", "Value"}
MX_REQUIRED = {"ScenarioID", "MessageType", "XPath", "Value"}

#: Messages the templates demonstrate out of the box.
MT_TEMPLATE_MESSAGES = ["MT541", "MT543", "MT548"]
MX_TEMPLATE_MESSAGES = ["sese.023", "sese.024", "sese.025"]


class ExcelFormatError(ValueError):
    """The workbook itself is unusable — wrong file type, missing columns, no rows."""


@dataclass
class ParsedScenario:
    scenario_id: str
    message_type: str
    profile_id: str
    row_numbers: list[int] = field(default_factory=list)
    fields: list[FieldInput] = field(default_factory=list)
    elements: list[ElementInput] = field(default_factory=list)
    issues: list[ValidationIssue] = field(default_factory=list)


@dataclass
class ParsedWorkbook:
    format: MessageFormat
    scenarios: list[ParsedScenario]


# --------------------------------------------------------------------------------------
# Template generation
# --------------------------------------------------------------------------------------


def _style_header(sheet: Worksheet, width_hint: dict[int, int] | None = None) -> None:
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.columns, start=1):
        letter = get_column_letter(index)
        longest = max((len(str(cell.value or "")) for cell in column), default=10)
        sheet.column_dimensions[letter].width = min(
            max((width_hint or {}).get(index, 0), longest + 2, 12), 70
        )


def _safe_cell(value: Any) -> Any:
    """Prefix formula-like text so Excel treats it as data, not a formula."""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def build_template(format_: MessageFormat, message_types: list[str] | None = None) -> bytes:
    targets = message_types or (
        MT_TEMPLATE_MESSAGES if format_ is MessageFormat.MT else MX_TEMPLATE_MESSAGES
    )
    workbook = Workbook()
    scenarios = workbook.active
    assert scenarios is not None
    scenarios.title = "Scenarios"
    headers = MT_HEADERS if format_ is MessageFormat.MT else MX_HEADERS
    scenarios.append(headers)

    for index, message_type in enumerate(targets, start=1):
        sample = build_sample(format_, message_type, SampleVariant.TYPICAL)
        scenario_id = f"TC{index:03d}"
        if format_ is MessageFormat.MT:
            for item in sample.inputs:
                scenarios.append(
                    [
                        scenario_id,
                        message_type,
                        "BASE_DEMO_V1",
                        item.sequence,
                        item.occurrence,
                        item.tag,
                        item.qualifier,
                        item.option,
                        _safe_cell(item.value),
                    ]
                )
        else:
            for element in sample.elements:
                scenarios.append(
                    [
                        scenario_id,
                        message_type,
                        "BASE_DEMO_V1",
                        element.path,
                        element.occurrence,
                        _safe_cell(element.value),
                    ]
                )
    _style_header(scenarios)
    _apply_code_dropdowns(scenarios, format_)

    # The Scenarios sheet stays a small, usable set of worked examples. The Reference
    # sheet is the field dictionary a tester actually looks things up in, so it covers
    # every configured message — otherwise a message is generatable but undiscoverable.
    _write_reference_sheet(workbook, format_, message_types or _all_message_types(format_))
    _write_codes_sheet(workbook, format_, message_types or _all_message_types(format_))
    _write_readme_sheet(workbook, format_)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _all_message_types(format_: MessageFormat) -> list[str]:
    from app.specifications.registry import specification_registry
    from app.studio.mx.registry import mx_registry

    if format_ is MessageFormat.MX:
        return [spec.message_type for spec in mx_registry.all_specs()]
    return [spec.message_type for spec in specification_registry.list()]


@dataclass(frozen=True)
class ReferenceRow:
    """One row of the template's Reference sheet: the field dictionary a tester looks in.

    Exposed as data rather than written straight into a worksheet so the coverage report can
    measure what the sheet really contains. That is not hypothetical: the sheet was once
    hardcoded to three MX messages while the registry held seven, and only reading it back
    would have caught it.
    """

    message_type: str
    field: SpecField


def reference_rows(
    format_: MessageFormat, message_types: list[str] | None = None
) -> list[ReferenceRow]:
    rows: list[ReferenceRow] = []
    for message_type in message_types or _all_message_types(format_):
        spec = message_spec(format_, message_type)
        rows.extend(
            ReferenceRow(message_type=spec.message_type, field=item)
            for item in sorted(spec.fields, key=lambda entry: entry.order)
        )
    return rows


def _write_reference_sheet(
    workbook: Workbook, format_: MessageFormat, targets: list[str]
) -> None:
    sheet = workbook.create_sheet("Reference")
    if format_ is MessageFormat.MT:
        sheet.append(
            [
                "MessageType",
                "Sequence",
                "Tag",
                "Qualifier",
                "Option",
                "Presence",
                "Field",
                "Input kind",
                "Format",
                "Example",
                "Literal written by the studio",
                "Max length",
                "Allowed codes",
            ]
        )
    else:
        sheet.append(
            [
                "MessageType",
                "XPath",
                "Presence",
                "Field",
                "Input kind",
                "Data type",
                "Format",
                "Example",
                "Allowed codes",
                "Max occurrences",
            ]
        )
    for entry in reference_rows(format_, targets):
        item = entry.field
        example = item.examples[0].value if item.examples else ""
        codes = ", ".join(item.allowed_codes)
        if format_ is MessageFormat.MT:
            row = [
                entry.message_type,
                item.sequence_code,
                item.tag,
                item.qualifier,
                item.option,
                item.presence.value,
                item.display_name,
                item.input_kind.value,
                item.format_explanation,
                _safe_cell(example),
                # Stated so an automation tester knows what NOT to type. 35B carries
                # "ISIN " here and nowhere in the Value column.
                item.literal_prefix or "",
                item.max_length or "",
                codes,
            ]
        else:
            row = [
                entry.message_type,
                item.xpath,
                item.presence.value,
                item.display_name,
                item.input_kind.value,
                item.data_type,
                item.format_explanation,
                _safe_cell(example),
                codes,
                item.max_occurs,
            ]
        sheet.append(row)
        if item.presence is Presence.MANDATORY:
            for cell in sheet[sheet.max_row]:
                cell.fill = MANDATORY_FILL
    _style_header(sheet)


#: Excel refuses an inline validation list longer than this, so a long code list gets a
#: pointer to the Codes sheet rather than a dropdown that silently fails to appear.
EXCEL_INLINE_LIST_LIMIT = 255


def _apply_code_dropdowns(sheet: Worksheet, format_: MessageFormat) -> None:
    """Put a real dropdown on the Value cell of every row whose field has a code list.

    The Scenarios sheet is one field per row, so the list has to be attached per cell rather
    than per column. Without this, a spreadsheet user is left guessing at TRAD, PAIR and
    TURN — which is exactly the position the browser used to leave a tester in.
    """
    if format_ is not MessageFormat.MT:
        return
    value_column = get_column_letter(len(MT_HEADERS))
    by_list: dict[str, list[str]] = {}
    for row_number in range(2, sheet.max_row + 1):
        message_type = sheet.cell(row_number, 1).value
        tag = sheet.cell(row_number, 6).value
        raw_qualifier = sheet.cell(row_number, 7).value
        if not message_type or not tag:
            continue
        qualifier = str(raw_qualifier) if raw_qualifier else None
        field = _lookup_field(str(message_type), str(tag), qualifier)
        if field is None or not field.allowed_codes:
            continue
        if field.input_kind is not InputKind.SELECT:
            # A code that is only part of the value — a quantity type such as UNIT/1000 —
            # would be broken by a whole-cell dropdown.
            continue
        formula = '"' + ",".join(field.allowed_codes) + '"'
        if len(formula) > EXCEL_INLINE_LIST_LIMIT:
            continue
        by_list.setdefault(formula, []).append(f"{value_column}{row_number}")
    for formula, cells in by_list.items():
        validation = DataValidation(
            type="list", formula1=formula, allow_blank=False, showDropDown=False
        )
        validation.error = "Choose one of the codes this field allows. See the Codes sheet."
        validation.errorTitle = "Not an allowed code"
        sheet.add_data_validation(validation)
        for cell in cells:
            validation.add(cell)


def _lookup_field(message_type: str, tag: str, qualifier: str | None) -> SpecField | None:
    try:
        spec = message_spec(MessageFormat.MT, message_type)
    except (KeyError, ValueError):
        return None
    return next(
        (
            item
            for item in spec.fields
            if item.tag == tag and (item.qualifier or None) == (qualifier or None)
        ),
        None,
    )


def _write_codes_sheet(workbook: Workbook, format_: MessageFormat, targets: list[str]) -> None:
    """Every controlled code, with the words for it.

    The same vocabulary the browser dropdown and the JSON API's `allowedValues` use, so a
    spreadsheet user and a UI user are never offered different codes for one field.
    """
    sheet = workbook.create_sheet("Codes")
    if format_ is MessageFormat.MT:
        sheet.append(
            [
                "MessageType",
                "Sequence",
                "Tag",
                "Qualifier",
                "Option",
                "Field",
                "Code",
                "Means",
                "Description",
            ]
        )
    else:
        sheet.append(["MessageType", "XPath", "Field", "Code", "Means", "Description"])
    for entry in reference_rows(format_, targets):
        item = entry.field
        for value in item.allowed_values:
            if format_ is MessageFormat.MT:
                sheet.append(
                    [
                        entry.message_type,
                        item.sequence_code,
                        item.tag,
                        item.qualifier,
                        item.option,
                        item.display_name,
                        _safe_cell(value.code),
                        value.label,
                        value.description,
                    ]
                )
            else:
                sheet.append(
                    [
                        entry.message_type,
                        item.xpath,
                        item.display_name,
                        _safe_cell(value.code),
                        value.label,
                        value.description,
                    ]
                )
    _style_header(sheet, width_hint={len(sheet[1]): 60})


def _write_readme_sheet(workbook: Workbook, format_: MessageFormat) -> None:
    sheet = workbook.create_sheet("Read me")
    sheet.append(["Column", "What it means"])
    common = [
        ("ScenarioID", "Groups rows into one message. All rows sharing an ID build one message."),
        ("MessageType", "The message to build, for example MT541 or sese.023."),
        (
            "ProfileID",
            "The client profile to validate against. Leave blank to use BASE_DEMO_V1.",
        ),
        ("Value", "The value to place in the field. See the Reference sheet for the format."),
    ]
    if format_ is MessageFormat.MT:
        rows = [
            *common[:3],
            (
                "Sequence",
                "The sequence code, for example GENL or SETDET. Leave blank when the tag "
                "appears in only one sequence.",
            ),
            (
                "SequenceOccurrence",
                "Which repeat of the sequence this row belongs to. Use 1 for the first, 2 for "
                "the second, and so on. Leave blank for 1.",
            ),
            ("Tag", "The field tag, for example 20C or 98A."),
            ("Qualifier", "The four-character qualifier, for example SEME or SETT."),
            ("Option", "Optional. The field option letter. It is checked against the format."),
            common[3],
        ]
    else:
        rows = [
            *common[:3],
            (
                "XPath",
                "The full element path, for example "
                "/Document/SctiesSttlmTxInstr/TxId. Copy these from the Reference sheet.",
            ),
            (
                "Occurrence",
                "Which repeat of a repeatable element this row is. Use 1 for the first. "
                "Leave blank for 1.",
            ),
            common[3],
        ]
    for row in rows:
        sheet.append(list(row))
    sheet.append([])
    sheet.append(["How to use this workbook", ""])
    for step in [
        "1. Edit the Scenarios sheet. Each ScenarioID becomes one generated message.",
        "2. Add rows for optional fields using the Reference sheet as your guide.",
        "3. Rows shaded in the Reference sheet are mandatory.",
        "4. Cells with a dropdown accept only the codes on the Codes sheet.",
        "5. Upload the workbook to POST /api/v1/messages/generate-from-excel.",
        "6. The response contains the generated message and per-scenario validation.",
    ]:
        sheet.append([step, ""])
    sheet.append([])
    sheet.append(["What goes in the Value column", ""])
    for note in [
        (
            "Business values, not SWIFT syntax",
            "Enter what the field means. The studio writes the SWIFT rendering.",
        ),
        (
            "Financial instrument (35B)",
            "Enter the 12-character identifier only, for example XS0000000009. Do NOT type "
            "the word ISIN — the studio writes it. A value that arrives with the prefix "
            "already on it is accepted and normalised, so an older sheet keeps working.",
        ),
        (
            "Settlement parties (95a)",
            "Use tag 95P and a BIC to identify a party by its Business Identifier Code. Use "
            "95R and a data source scheme with the identifier, such as CSD/DEMOPSET01, for "
            "a proprietary code. Do not put a BIC in a 95R value.",
        ),
        (
            "Controlled codes",
            "Use the Code column on the Codes sheet. The Means column is there to read, not "
            "to enter.",
        ),
        (
            "Same values, same message",
            "This workbook, the JSON API and the browser produce byte-identical output for "
            "equivalent values, because all three call the same composer.",
        ),
    ]:
        sheet.append(list(note))
    for cell in sheet["A"]:
        # Rows with no second column are section headings, so make them stand out.
        if cell.row > 1 and not sheet.cell(cell.row, 2).value:
            cell.font = GROUP_FONT
    _style_header(sheet, width_hint={1: 26, 2: 90})
    sheet.column_dimensions["B"].width = 90
    for row in sheet.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


# --------------------------------------------------------------------------------------
# Parsing
# --------------------------------------------------------------------------------------


def _text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        # Excel silently turns 2026-08-18 into a datetime; MX wants the ISO date back.
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _occurrence(value: Any) -> int:
    text = _text(value)
    if not text:
        return 1
    try:
        parsed = int(float(text))
    except ValueError:
        return 1
    return parsed if 1 <= parsed <= 100 else 1


def _row_issue(row_number: int, message: str, suggestion: str) -> ValidationIssue:
    return ValidationIssue(
        rule_id="EXCEL_ROW_INVALID",
        severity=IssueSeverity.ERROR,
        layer=ValidationLayer.CANONICAL,
        location=f"row {row_number}",
        message=message,
        suggestion=suggestion,
    )


def validate_upload(content: bytes, filename: str, max_bytes: int) -> None:
    if Path(filename).name != filename or ".." in filename:
        raise ExcelFormatError("The upload filename contains a path component.")
    if Path(filename).suffix.lower() != ".xlsx":
        raise ExcelFormatError("Only .xlsx workbooks are supported.")
    if len(content) > max_bytes:
        raise ExcelFormatError(f"The workbook exceeds the {max_bytes}-byte upload limit.")
    if not content.startswith(b"PK") or not is_zipfile(BytesIO(content)):
        raise ExcelFormatError("The upload is not a valid .xlsx workbook.")


def parse_workbook(
    content: bytes,
    *,
    format_: MessageFormat | None = None,
    default_profile_id: str = "BASE_DEMO_V1",
    max_rows: int = 5_000,
) -> ParsedWorkbook:
    """Read the Scenarios sheet into one :class:`ParsedScenario` per ScenarioID."""
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["Scenarios"] if "Scenarios" in workbook.sheetnames else workbook.active
    if sheet is None:
        raise ExcelFormatError("The workbook has no readable sheet.")
    rows = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as error:
        raise ExcelFormatError("The workbook is empty.") from error

    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    normalised = [re.sub(r"[^a-z0-9]", "", header.lower()) for header in headers]
    if len(normalised) != len(set(normalised)):
        raise ExcelFormatError("The Scenarios sheet has duplicate column headers.")

    detected = format_ or _detect_format(normalised)
    required = MT_REQUIRED if detected is MessageFormat.MT else MX_REQUIRED
    expected = MT_HEADERS if detected is MessageFormat.MT else MX_HEADERS
    lookup = {
        re.sub(r"[^a-z0-9]", "", header.lower()): index
        for index, header in enumerate(headers)
    }
    missing = [
        header
        for header in required
        if re.sub(r"[^a-z0-9]", "", header.lower()) not in lookup
    ]
    if missing:
        raise ExcelFormatError(
            f"The {detected.value} sheet is missing required columns: "
            f"{', '.join(sorted(missing))}. Expected columns: {', '.join(expected)}."
        )

    def cell(row: tuple[Any, ...], header: str) -> Any:
        index = lookup.get(re.sub(r"[^a-z0-9]", "", header.lower()))
        return row[index] if index is not None and index < len(row) else None

    scenarios: dict[str, ParsedScenario] = {}
    order: list[str] = []
    count = 0
    for row_number, raw in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in raw):
            continue
        count += 1
        if count > max_rows:
            raise ExcelFormatError(
                f"The workbook has more than {max_rows} data rows."
            )
        scenario_id = _text(cell(raw, "ScenarioID"))
        message_type = _text(cell(raw, "MessageType"))
        value = _text(cell(raw, "Value"))
        profile_id = _text(cell(raw, "ProfileID")) or default_profile_id

        if not scenario_id:
            scenario_id = "UNASSIGNED"
        if scenario_id not in scenarios:
            scenarios[scenario_id] = ParsedScenario(
                scenario_id=scenario_id,
                message_type=message_type or "",
                profile_id=profile_id,
            )
            order.append(scenario_id)
        scenario = scenarios[scenario_id]
        scenario.row_numbers.append(row_number)

        if message_type and not scenario.message_type:
            scenario.message_type = message_type
        elif message_type and message_type.lower() != scenario.message_type.lower():
            scenario.issues.append(
                _row_issue(
                    row_number,
                    f"Scenario {scenario_id} mixes message types "
                    f"{scenario.message_type} and {message_type}.",
                    "Give each message its own ScenarioID.",
                )
            )
            continue
        if not scenario.message_type:
            scenario.issues.append(
                _row_issue(
                    row_number,
                    "MessageType is missing.",
                    "Add the message type, for example MT541 or sese.023.",
                )
            )
            continue
        if value is None:
            scenario.issues.append(
                _row_issue(
                    row_number,
                    "Value is empty.",
                    "Enter a value, or delete the row if the field is not needed.",
                )
            )
            continue

        if detected is MessageFormat.MT:
            tag = _text(cell(raw, "Tag"))
            if not tag:
                scenario.issues.append(
                    _row_issue(
                        row_number,
                        "Tag is missing.",
                        "Add the field tag, for example 20C. The Reference sheet lists them.",
                    )
                )
                continue
            scenario.fields.append(
                FieldInput(
                    sequence=_text(cell(raw, "Sequence")),
                    occurrence=_occurrence(cell(raw, "SequenceOccurrence")),
                    tag=tag,
                    qualifier=_text(cell(raw, "Qualifier")),
                    option=_text(cell(raw, "Option")),
                    value=value,
                )
            )
        else:
            xpath = _text(cell(raw, "XPath"))
            if not xpath:
                scenario.issues.append(
                    _row_issue(
                        row_number,
                        "XPath is missing.",
                        "Add the element path. The Reference sheet lists every supported path.",
                    )
                )
                continue
            scenario.elements.append(
                ElementInput(
                    path=xpath,
                    occurrence=_occurrence(cell(raw, "Occurrence")),
                    value=value,
                )
            )
    if not order:
        raise ExcelFormatError("The Scenarios sheet has no data rows.")
    return ParsedWorkbook(format=detected, scenarios=[scenarios[key] for key in order])


def _detect_format(normalised_headers: list[str]) -> MessageFormat:
    if "xpath" in normalised_headers:
        return MessageFormat.MX
    if "tag" in normalised_headers:
        return MessageFormat.MT
    raise ExcelFormatError(
        "The sheet has neither a Tag column (MT) nor an XPath column (MX). "
        "Download a template to see the expected layout."
    )
