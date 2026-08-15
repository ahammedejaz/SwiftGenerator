import json
import re
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, is_zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from app.config import Settings
from app.domain.enums import (
    GenerationMode,
    Lifecycle,
    NegativeMutation,
    ResponseAction,
    StatusCategory,
)
from app.domain.models import (
    Account,
    BulkGenerateResponse,
    BulkRowResult,
    GeneratedMessage,
    LifecycleResponseRequest,
    Security,
    Settlement,
    SettlementScenario,
    TestConfiguration,
    Trade,
    ValidationFinding,
)
from app.persistence.reports import ReportRepository
from app.services.generation import DISCLAIMER, DomainValidationError, GenerationService
from app.services.lifecycle import LifecycleService

REQUIRED_HEADERS = [
    "Scenario ID",
    "Profile ID",
    "Lifecycle",
    "Direction",
    "Payment Type",
    "Function",
    "Sender Reference",
    "Related Reference",
    "Transaction Type",
    "ISIN",
    "Quantity",
    "Trade Date",
    "Settlement Date",
    "Safekeeping Account",
    "Place of Settlement",
    "Delivering Agent",
    "Receiving Agent",
    "Currency",
    "Amount",
    "Status Category",
    "Status Code",
    "Reason Code",
    "Reason Narrative",
    "Generation Mode",
    "Negative Mutation",
]
OPTIONAL_HEADERS = ["Client Reference", "Settlement Result", "Actual Settlement Date"]
TEMPLATE_HEADERS = [*REQUIRED_HEADERS, *OPTIONAL_HEADERS]

STATUS_ACTIONS = {
    StatusCategory.PENDING: ResponseAction.PENDING_STATUS,
    StatusCategory.REJECTED: ResponseAction.REJECTED_STATUS,
    StatusCategory.MATCHED: ResponseAction.MATCHED_STATUS,
    StatusCategory.UNMATCHED: ResponseAction.UNMATCHED_STATUS,
    StatusCategory.CANCELLATION_ACCEPTED: ResponseAction.CANCELLATION_ACCEPTED_STATUS,
    StatusCategory.CANCELLATION_REJECTED: ResponseAction.CANCELLATION_REJECTED_STATUS,
}


class BulkService:
    def __init__(
        self,
        settings: Settings,
        generation: GenerationService,
        lifecycle: LifecycleService,
        reports: ReportRepository,
    ) -> None:
        self._settings = settings
        self._generation = generation
        self._lifecycle = lifecycle
        self._reports = reports

    def template(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Scenarios"
        sheet.append(TEMPLATE_HEADERS)
        samples = [
            self._sample_row("BULK-MT540", "RECEIVE", "FREE_OF_PAYMENT", "BUY"),
            self._sample_row("BULK-MT541", "RECEIVE", "AGAINST_PAYMENT", "BUY"),
            self._sample_row("BULK-MT543", "DELIVER", "AGAINST_PAYMENT", "SELL"),
            self._sample_row(
                "BULK-INVALID-MT541", "RECEIVE", "AGAINST_PAYMENT", "BUY", amount=None
            ),
        ]
        for row in samples:
            sheet.append([row.get(header) for header in TEMPLATE_HEADERS])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="006D77")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(sheet.columns, start=1):
            letter = get_column_letter(index)
            sheet.column_dimensions[letter].width = min(
                max(len(str(cell.value or "")) for cell in column) + 2, 32
            )
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def generate(self, content: bytes, filename: str) -> BulkGenerateResponse:
        self._validate_file(content, filename)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        assert sheet is not None
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise ValueError("The workbook is empty") from exc
        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
        if missing_headers:
            raise ValueError(f"Missing mandatory Excel columns: {', '.join(missing_headers)}")
        if len(headers) != len(set(headers)):
            raise ValueError("The workbook contains duplicate column headers")

        data_rows = list(rows)
        if len(data_rows) > self._settings.max_bulk_rows:
            raise ValueError(
                f"Workbook contains {len(data_rows)} rows; "
                f"maximum is {self._settings.max_bulk_rows}"
            )

        results: list[BulkRowResult] = []
        generated_messages = []
        instruction_by_reference: dict[str, str] = {}
        for row_number, raw_row in enumerate(data_rows, start=2):
            values = dict(zip(headers, raw_row, strict=False))
            if not any(value not in (None, "") for value in raw_row):
                continue
            scenario_id = self._text(values.get("Scenario ID")) or f"ROW-{row_number}"
            try:
                lifecycle = Lifecycle(self._enum_text(values.get("Lifecycle")))
                if lifecycle == Lifecycle.INSTRUCTION:
                    scenario = self._instruction_scenario(values, scenario_id)
                    generated = self._generation.generate(scenario)
                    if generated.scenario.sender_reference:
                        instruction_by_reference[generated.scenario.sender_reference] = (
                            generated.message_id
                        )
                else:
                    generated = self._generate_response(
                        values, scenario_id, lifecycle, instruction_by_reference
                    )
                generated_messages.append(generated)
                filename_stem = self._safe_stem(scenario_id)
                generated_filename = (
                    f"{row_number:04d}_{filename_stem}_{generated.resolved_message_type.value}.txt"
                )
                results.append(
                    BulkRowResult(
                        row_number=row_number,
                        scenario_id=scenario_id,
                        status="GENERATED",
                        resolved_message_type=generated.resolved_message_type,
                        message_id=generated.message_id,
                        generated_filename=generated_filename,
                        profile_id=generated.profile_id,
                        profile_version=generated.profile_version,
                        validation_status=generated.validation.status,
                        error_count=generated.validation.error_count,
                        warning_count=generated.validation.warning_count,
                        expected_negative_failure=(
                            generated.validation.status.value == "INTENTIONALLY_INVALID"
                        ),
                        findings=generated.validation.findings,
                    )
                )
            except (ValueError, KeyError, ValidationError, DomainValidationError) as exc:
                findings = (
                    exc.report.findings
                    if isinstance(exc, DomainValidationError)
                    else [self._row_finding(str(exc))]
                )
                results.append(
                    BulkRowResult(
                        row_number=row_number,
                        scenario_id=scenario_id,
                        status="FAILED",
                        error_count=max(1, len(findings)),
                        findings=findings,
                    )
                )

        generated_count = sum(item.status == "GENERATED" for item in results)
        failed_count = sum(item.status == "FAILED" for item in results)
        report_payload = {
            "totalRows": len(results),
            "generatedRows": generated_count,
            "failedRows": failed_count,
            "disclaimer": DISCLAIMER,
            "rows": [item.model_dump(mode="json", by_alias=True) for item in results],
        }
        artifact = self._build_zip(generated_messages, results, report_payload)
        report_id = self._reports.save_zip(artifact, report_payload)
        return BulkGenerateResponse(
            report_id=report_id,
            total_rows=len(results),
            generated_rows=generated_count,
            failed_rows=failed_count,
            row_results=results,
            download_path=f"/api/reports/{report_id}",
            disclaimer=DISCLAIMER,
        )

    def _instruction_scenario(self, values: dict[str, Any], scenario_id: str) -> SettlementScenario:
        payment_type = self._enum_text(values.get("Payment Type"))
        mode = GenerationMode(
            self._enum_text(values.get("Generation Mode") or GenerationMode.VALID.value)
        )
        mutation_value = self._text(values.get("Negative Mutation"))
        return SettlementScenario(
            scenario_id=scenario_id,
            profile_id=self._text(values.get("Profile ID")) or "BASE_DEMO_V1",
            lifecycle=Lifecycle.INSTRUCTION,
            direction=self._enum_text(values.get("Direction")),
            payment_type=payment_type,
            function=self._enum_text(values.get("Function") or "NEWM"),
            sender_reference=self._text(values.get("Sender Reference")),
            related_reference=self._text(values.get("Related Reference")),
            client_reference=self._text(values.get("Client Reference")),
            trade=Trade(
                transaction_type=self._enum_text(values.get("Transaction Type")),
                trade_date=self._date(values.get("Trade Date")),
                settlement_date=self._date(values.get("Settlement Date")),
            ),
            security=Security(
                identifier=self._text(values.get("ISIN")),
                quantity=self._decimal(values.get("Quantity")),
            ),
            account=Account(safekeeping_account=self._text(values.get("Safekeeping Account"))),
            settlement=Settlement(
                currency=self._text(values.get("Currency")),
                amount=self._decimal(values.get("Amount")),
                place_of_settlement=self._text(values.get("Place of Settlement")),
                delivering_agent=self._text(values.get("Delivering Agent")),
                receiving_agent=self._text(values.get("Receiving Agent")),
            ),
            test_configuration=TestConfiguration(
                mode=mode,
                mutation=NegativeMutation(self._enum_text(mutation_value))
                if mutation_value
                else None,
            ),
            synthetic_data=True,
        )

    def _generate_response(
        self,
        values: dict[str, Any],
        scenario_id: str,
        lifecycle: Lifecycle,
        instructions: dict[str, str],
    ) -> GeneratedMessage:
        del scenario_id
        related_reference = self._text(values.get("Related Reference"))
        if not related_reference or related_reference not in instructions:
            raise ValueError("Confirmation or status row must reference an earlier instruction row")
        mode = GenerationMode(
            self._enum_text(values.get("Generation Mode") or GenerationMode.VALID.value)
        )
        mutation_value = self._text(values.get("Negative Mutation"))
        if lifecycle == Lifecycle.STATUS:
            category = StatusCategory(self._enum_text(values.get("Status Category")))
            action = STATUS_ACTIONS[category]
        else:
            settlement_result = self._enum_text(values.get("Settlement Result") or "FULL")
            action = (
                ResponseAction.PARTIAL_CONFIRMATION
                if settlement_result == "PARTIAL"
                else ResponseAction.FULL_CONFIRMATION
            )
        request = LifecycleResponseRequest(
            action=action,
            response_reference=self._text(values.get("Sender Reference")),
            reason_code=self._text(values.get("Reason Code")),
            reason_narrative=self._text(values.get("Reason Narrative")),
            actual_settlement_date=self._date(
                values.get("Actual Settlement Date") or values.get("Settlement Date")
            ),
            settled_quantity=self._decimal(values.get("Quantity")),
            settled_amount=self._decimal(values.get("Amount")),
            generation_mode=mode,
            negative_mutation=(
                NegativeMutation(self._enum_text(mutation_value)) if mutation_value else None
            ),
        )
        return self._lifecycle.generate_response(instructions[related_reference], request)

    def _build_zip(self, messages, results, report_payload) -> bytes:  # type: ignore[no-untyped-def]
        output = BytesIO()
        message_by_id = {message.message_id: message for message in messages}
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
            for result in results:
                if not result.message_id or not result.generated_filename:
                    continue
                message = message_by_id[result.message_id]
                archive.writestr(result.generated_filename, message.raw_message)
                report_name = result.generated_filename.removesuffix(".txt") + ".validation.json"
                archive.writestr(
                    report_name,
                    json.dumps(
                        message.validation.model_dump(mode="json", by_alias=True),
                        indent=2,
                    ),
                )
            archive.writestr("summary.xlsx", self._summary_workbook(results))
            archive.writestr("execution-report.json", json.dumps(report_payload, indent=2))
        return output.getvalue()

    def _summary_workbook(self, results: list[BulkRowResult]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Execution Summary"
        headers = [
            "Row",
            "Scenario ID",
            "Resolved Message Type",
            "Generated Filename",
            "Profile ID",
            "Profile Version",
            "Validation Result",
            "Error Count",
            "Warning Count",
            "Expected Negative Failure",
            "Actual Outcome",
        ]
        sheet.append(headers)
        for result in results:
            values = [
                result.row_number,
                result.scenario_id,
                result.resolved_message_type.value if result.resolved_message_type else None,
                result.generated_filename,
                result.profile_id,
                result.profile_version,
                result.validation_status.value if result.validation_status else "FAILED",
                result.error_count,
                result.warning_count,
                result.expected_negative_failure,
                result.status,
            ]
            sheet.append([self._safe_excel_cell(value) for value in values])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _validate_file(self, content: bytes, filename: str) -> None:
        if Path(filename).name != filename or ".." in filename:
            raise ValueError("Upload filename contains a path component")
        if Path(filename).suffix.lower() != ".xlsx":
            raise ValueError("Only .xlsx workbooks are supported")
        if len(content) > self._settings.max_upload_bytes:
            raise ValueError(
                f"Workbook exceeds the {self._settings.max_upload_bytes}-byte upload limit"
            )
        if not content.startswith(b"PK") or not is_zipfile(BytesIO(content)):
            raise ValueError("Upload is not a valid OOXML workbook")

    @staticmethod
    def _sample_row(
        scenario_id: str,
        direction: str,
        payment_type: str,
        transaction_type: str,
        amount: str | None = "25000.00",
    ) -> dict[str, Any]:
        is_fop = payment_type == "FREE_OF_PAYMENT"
        return {
            "Scenario ID": scenario_id,
            "Profile ID": "BASE_DEMO_V1",
            "Lifecycle": "INSTRUCTION",
            "Direction": direction,
            "Payment Type": payment_type,
            "Function": "NEWM",
            "Sender Reference": scenario_id.replace("-", "")[:16],
            "Transaction Type": transaction_type,
            "ISIN": "XS0000000001",
            "Quantity": "1000",
            "Trade Date": "2026-08-03",
            "Settlement Date": "2026-08-06",
            "Safekeeping Account": "SYNTHSAFE01",
            "Place of Settlement": "SYNTHPSET01",
            "Delivering Agent": "SYNTHDEAG01",
            "Receiving Agent": "SYNTHREAG01",
            "Currency": None if is_fop else "USD",
            "Amount": None if is_fop else amount,
            "Generation Mode": "VALID",
        }

    @staticmethod
    def _enum_text(value: Any) -> str:
        return str(value or "").strip().upper().replace(" ", "_")

    @staticmethod
    def _text(value: Any) -> str | None:
        if value in (None, ""):
            return None
        return str(value).strip()

    @staticmethod
    def _decimal(value: Any) -> Decimal | None:
        if value in (None, ""):
            return None
        return Decimal(str(value).replace(",", ""))

    @staticmethod
    def _date(value: Any) -> date | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value).strip())

    @staticmethod
    def _safe_stem(value: str) -> str:
        cleaned = re.sub(r"[^A-Za-z0-9_-]", "_", value).strip("._")
        return cleaned[:64] or "scenario"

    @staticmethod
    def _safe_excel_cell(value: Any) -> Any:
        if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
            return f"'{value}"
        return value

    @staticmethod
    def _row_finding(message: str) -> ValidationFinding:
        return ValidationFinding(
            rule_id="BULK-ROW-INVALID",
            severity="ERROR",
            field_path=None,
            message="The Excel row could not be generated.",
            technical_explanation=message,
            expected_condition="A supported canonical scenario",
            suggestion="Correct the row values and upload the workbook again.",
        )
