from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, is_zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import ValidationError

from app.config import Settings
from app.domain.enums import SettlementCommandType
from app.domain.models import (
    BulkGenerateResponse,
    BulkRowResult,
    SettlementCommandRequest,
    ValidationFinding,
)
from app.persistence.reports import ReportRepository
from app.services.generation import DISCLAIMER
from app.workflows.corporate_actions import (
    CorporateActionConfirmationRequest,
    CorporateActionInstructionRequest,
    CorporateActionNarrativeRequest,
    CorporateActionNotification,
    CorporateActionNotificationRequest,
    CorporateActionStatusRequest,
    CorporateActionWorkflowService,
)
from app.workflows.penalties import (
    PenaltyEntry,
    PenaltyGenerateRequest,
    PenaltyStatement,
    PenaltyWorkflowService,
)
from app.workflows.settlement_processing import SettlementProcessingService

HEADERS = [
    "Workflow Type",
    "Scenario ID",
    "Profile ID",
    "Workflow ID",
    "Message Reference",
    "Related Message Reference",
    "Original Instruction ID",
    "Event Reference",
    "Event Type",
    "Security",
    "Safekeeping Account",
    "Eligible Quantity",
    "Election Deadline",
    "Payment Date",
    "Option Number",
    "Option Code",
    "Default Option",
    "Quantity",
    "Status",
    "Reason Code",
    "Currency",
    "Amount",
    "Narrative",
    "Penalty Reference",
    "Related Instruction Reference",
    "Penalty Type",
    "Penalty Action",
    "Amount Direction",
    "Detection Date",
    "Number Of Days",
    "Command Type",
    "Priority",
]
REQUIRED_HEADERS = HEADERS[:6]


class WorkflowBulkService:
    """Row-isolated Excel support for implemented non-settlement workflow modules."""

    def __init__(
        self,
        settings: Settings,
        reports: ReportRepository,
        settlement: SettlementProcessingService,
        penalties: PenaltyWorkflowService,
        corporate_actions: CorporateActionWorkflowService,
    ) -> None:
        self._settings = settings
        self._reports = reports
        self._settlement = settlement
        self._penalties = penalties
        self._corporate_actions = corporate_actions

    def template(self) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Workflow Scenarios"
        sheet.append(HEADERS)
        samples: list[dict[str, Any]] = [
            {
                "Workflow Type": "PENALTY",
                "Scenario ID": "SYNTH-PENA-1",
                "Profile ID": "BASE_DEMO_V1",
                "Workflow ID": "SYNTH-PENA-WF",
                "Message Reference": "PENASTMT0001",
                "Safekeeping Account": "SYNTHSAFE01",
                "Status": "ACTIVE",
                "Currency": "EUR",
                "Amount": "25.00",
                "Penalty Reference": "PENALTY0001",
                "Related Instruction Reference": "SYNTHSETTLE01",
                "Penalty Type": "SETTLEMENT_FAIL",
                "Penalty Action": "NEW",
                "Amount Direction": "PAYABLE",
                "Detection Date": "2026-08-04",
                "Number Of Days": 1,
            },
            {
                "Workflow Type": "CORPORATE_NOTIFICATION",
                "Scenario ID": "SYNTH-CA-1",
                "Profile ID": "BASE_DEMO_V1",
                "Workflow ID": "SYNTH-CA-WF",
                "Message Reference": "CA564SYNTH001",
                "Event Reference": "CAEVSYNTH001",
                "Event Type": "DIVIDEND_WITH_OPTIONS",
                "Security": "XS0000000001",
                "Safekeeping Account": "SYNTHSAFE01",
                "Eligible Quantity": "1000",
                "Election Deadline": "2099-08-10",
                "Payment Date": "2099-08-15",
            },
        ]
        for sample in samples:
            sheet.append([sample.get(header) for header in HEADERS])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="006D77")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def generate(self, content: bytes, filename: str) -> BulkGenerateResponse:
        self._validate_file(content, filename)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        assert sheet is not None
        rows = sheet.iter_rows(values_only=True)
        try:
            headers = [self._text(item) or "" for item in next(rows)]
        except StopIteration as exc:
            raise ValueError("The workbook is empty") from exc
        missing = [header for header in REQUIRED_HEADERS if header not in headers]
        if missing:
            raise ValueError(f"Missing mandatory Excel columns: {', '.join(missing)}")
        if len(headers) != len(set(headers)):
            raise ValueError("The workbook contains duplicate column headers")
        data_rows = list(rows)
        if len(data_rows) > self._settings.max_bulk_rows:
            raise ValueError("Workflow workbook exceeds the configured row limit")

        results: list[BulkRowResult] = []
        generated: list[Any] = []
        by_reference: dict[str, Any] = {}
        for row_number, row in enumerate(data_rows, start=2):
            values = dict(zip(headers, row, strict=False))
            if not any(value not in (None, "") for value in row):
                continue
            scenario_id = self._text(values.get("Scenario ID")) or f"ROW-{row_number}"
            workflow_type = (self._text(values.get("Workflow Type")) or "").upper()
            try:
                message = self._generate_row(workflow_type, values, by_reference)
                generated.append(message)
                canonical = message.canonical_data if hasattr(message, "canonical_data") else {}
                message_reference = self._text(values.get("Message Reference"))
                if message_reference:
                    by_reference[message_reference] = message
                result_filename = (
                    f"{row_number:04d}_{self._safe_stem(scenario_id)}_"
                    f"{message.resolved_message_type.value}.txt"
                )
                results.append(
                    BulkRowResult(
                        row_number=row_number,
                        scenario_id=scenario_id,
                        status="GENERATED",
                        resolved_message_type=message.resolved_message_type,
                        message_id=message.message_id,
                        generated_filename=result_filename,
                        profile_id=message.profile_id,
                        profile_version=message.profile_version,
                        validation_status=message.validation.status,
                        error_count=message.validation.error_count,
                        warning_count=message.validation.warning_count,
                        expected_negative_failure=False,
                        findings=message.validation.findings,
                    )
                )
                del canonical
            except (ValueError, KeyError, ValidationError) as exc:
                results.append(
                    BulkRowResult(
                        row_number=row_number,
                        scenario_id=scenario_id,
                        status="FAILED",
                        error_count=1,
                        findings=[self._finding(str(exc))],
                    )
                )

        report_payload = {
            "workflowBulk": True,
            "totalRows": len(results),
            "generatedRows": sum(item.status == "GENERATED" for item in results),
            "failedRows": sum(item.status == "FAILED" for item in results),
            "rows": [item.model_dump(mode="json", by_alias=True) for item in results],
            "disclaimer": DISCLAIMER,
        }
        artifact = self._zip(generated, results, report_payload)
        report_id = self._reports.save_zip(artifact, report_payload)
        return BulkGenerateResponse(
            report_id=report_id,
            total_rows=len(results),
            generated_rows=report_payload["generatedRows"],
            failed_rows=report_payload["failedRows"],
            row_results=results,
            download_path=f"/api/reports/{report_id}",
            disclaimer=DISCLAIMER,
        )

    def _generate_row(
        self, workflow_type: str, values: dict[str, Any], by_reference: dict[str, Any]
    ) -> Any:
        profile_id = self._text(values.get("Profile ID")) or "BASE_DEMO_V1"
        workflow_id = self._required(values, "Workflow ID")
        reference = self._required(values, "Message Reference")
        related_reference = self._text(values.get("Related Message Reference"))
        related = by_reference.get(related_reference or "")
        if workflow_type == "SETTLEMENT_COMMAND":
            return self._settlement.command(
                SettlementCommandRequest(
                    original_instruction_id=self._required(values, "Original Instruction ID"),
                    command_reference=reference,
                    command_type=SettlementCommandType(
                        self._text(values.get("Command Type")) or "MODIFY_PRIORITY"
                    ),
                    priority=int(self._required(values, "Priority")),
                )
            )
        if workflow_type == "PENALTY":
            detection = self._date(values.get("Detection Date"))
            entry = PenaltyEntry(
                penalty_reference=self._required(values, "Penalty Reference"),
                related_instruction_reference=self._text(
                    values.get("Related Instruction Reference")
                ),
                penalty_type=self._required(values, "Penalty Type"),
                action=self._text(values.get("Penalty Action")) or "NEW",
                status=self._text(values.get("Status")) or "ACTIVE",
                currency=self._required(values, "Currency"),
                amount=self._decimal(values.get("Amount")),
                amount_direction=self._required(values, "Amount Direction"),
                detection_date=detection,
                number_of_days=int(self._text(values.get("Number Of Days")) or "0"),
            )
            return self._penalties.generate(
                PenaltyGenerateRequest(
                    statement=PenaltyStatement(
                        workflow_id=workflow_id,
                        profile_id=profile_id,
                        statement_reference=reference,
                        statement_date=detection,
                        safekeeping_account=self._required(values, "Safekeeping Account"),
                        account_servicer="SYNTHSERVICER",
                        related_party="SYNTHPARTY",
                        list_type="NEW_ONLY",
                        penalties=[entry],
                    )
                )
            )
        if workflow_type == "CORPORATE_NOTIFICATION":
            return self._corporate_actions.notification(
                CorporateActionNotificationRequest(
                    notification=CorporateActionNotification(
                        workflow_id=workflow_id,
                        profile_id=profile_id,
                        event_reference=self._required(values, "Event Reference"),
                        message_reference=reference,
                        event_type=self._text(values.get("Event Type")) or "DIVIDEND_WITH_OPTIONS",
                        security_identifier=self._required(values, "Security"),
                        safekeeping_account=self._required(values, "Safekeeping Account"),
                        eligible_quantity=self._decimal(values.get("Eligible Quantity")),
                        election_deadline=self._date(values.get("Election Deadline")),
                        payment_date=self._date(values.get("Payment Date")),
                        options=[
                            {"optionNumber": 1, "optionCode": "CASH", "defaultOption": True},
                            {
                                "optionNumber": 2,
                                "optionCode": "SECURITIES",
                                "defaultOption": False,
                            },
                        ],
                    )
                )
            )
        if related is None:
            raise ValueError("Related Message Reference must identify an earlier valid row")
        if workflow_type == "CORPORATE_INSTRUCTION":
            return self._corporate_actions.instruction(
                CorporateActionInstructionRequest(
                    workflow_id=workflow_id,
                    profile_id=profile_id,
                    message_reference=reference,
                    notification_message_id=related.message_id,
                    option_number=int(self._text(values.get("Option Number")) or "1"),
                    instructed_quantity=self._decimal(values.get("Quantity")),
                )
            )
        if workflow_type == "CORPORATE_STATUS":
            return self._corporate_actions.status(
                CorporateActionStatusRequest(
                    workflow_id=workflow_id,
                    profile_id=profile_id,
                    message_reference=reference,
                    instruction_message_id=related.message_id,
                    status=self._required(values, "Status"),
                    reason_code=self._text(values.get("Reason Code")),
                )
            )
        if workflow_type == "CORPORATE_CONFIRMATION":
            return self._corporate_actions.confirmation(
                CorporateActionConfirmationRequest(
                    workflow_id=workflow_id,
                    profile_id=profile_id,
                    message_reference=reference,
                    instruction_message_id=related.message_id,
                    option_number=int(self._text(values.get("Option Number")) or "1"),
                    confirmed_quantity=self._decimal(values.get("Quantity")),
                    cash_currency=self._text(values.get("Currency")),
                    cash_amount=self._decimal(values.get("Amount")),
                    payment_date=self._date(values.get("Payment Date")),
                )
            )
        if workflow_type == "CORPORATE_NARRATIVE":
            return self._corporate_actions.narrative(
                CorporateActionNarrativeRequest(
                    workflow_id=workflow_id,
                    profile_id=profile_id,
                    message_reference=reference,
                    notification_message_id=related.message_id,
                    narrative=self._required(values, "Narrative"),
                )
            )
        raise ValueError(f"Unsupported workflow type: {workflow_type or 'blank'}")

    def _zip(
        self, generated: list[Any], results: list[BulkRowResult], payload: dict[str, Any]
    ) -> bytes:
        output = BytesIO()
        by_id = {message.message_id: message for message in generated}
        with ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
            for result in results:
                if not result.message_id or not result.generated_filename:
                    continue
                message = by_id[result.message_id]
                archive.writestr(result.generated_filename, message.raw_message)
                archive.writestr(
                    result.generated_filename.removesuffix(".txt") + ".validation.json",
                    json.dumps(message.validation.model_dump(mode="json", by_alias=True), indent=2),
                )
            archive.writestr("execution-report.json", json.dumps(payload, indent=2))
            archive.writestr("summary.xlsx", self._summary(results))
        return output.getvalue()

    def _summary(self, results: list[BulkRowResult]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "Workflow Summary"
        sheet.append(["Row", "Scenario ID", "Message Type", "Result", "Profile", "Errors"])
        for item in results:
            values = [
                item.row_number,
                item.scenario_id,
                item.resolved_message_type.value if item.resolved_message_type else None,
                item.status,
                item.profile_id,
                item.error_count,
            ]
            sheet.append([self._safe_excel(value) for value in values])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _validate_file(self, content: bytes, filename: str) -> None:
        if Path(filename).name != filename or ".." in filename:
            raise ValueError("Upload filename contains a path component")
        if Path(filename).suffix.lower() != ".xlsx":
            raise ValueError("Only .xlsx workbooks are supported")
        if len(content) > self._settings.max_upload_bytes:
            raise ValueError("Workbook exceeds the configured upload limit")
        if not is_zipfile(BytesIO(content)):
            raise ValueError("Uploaded content is not a valid .xlsx archive")

    @staticmethod
    def _required(values: dict[str, Any], name: str) -> str:
        value = WorkflowBulkService._text(values.get(name))
        if not value:
            raise ValueError(f"{name} is required")
        return value

    @staticmethod
    def _text(value: Any) -> str | None:
        if value in (None, ""):
            return None
        return str(value).strip()

    @staticmethod
    def _decimal(value: Any) -> Decimal:
        if value in (None, ""):
            raise ValueError("A required numeric value is missing")
        return Decimal(str(value))

    @staticmethod
    def _date(value: Any) -> date:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = WorkflowBulkService._text(value)
        if not text:
            raise ValueError("A required date is missing")
        return date.fromisoformat(text)

    @staticmethod
    def _safe_stem(value: str) -> str:
        cleaned = "".join(character if character.isalnum() else "_" for character in value)
        return cleaned[:64] or "scenario"

    @staticmethod
    def _safe_excel(value: Any) -> Any:
        if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
            return "'" + value
        return value

    @staticmethod
    def _finding(message: str) -> ValidationFinding:
        return ValidationFinding(
            rule_id="WORKFLOW-BULK-ROW",
            severity="ERROR",
            field_path=None,
            message=message,
            technical_explanation="The row failed typed deterministic workflow validation.",
            expected_condition="Provide all fields required for the selected workflow type.",
            suggestion="Correct this row and import it again.",
        )
