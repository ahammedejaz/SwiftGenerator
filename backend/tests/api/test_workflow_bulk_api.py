from io import BytesIO
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

from app.bulk.workflow_service import HEADERS


def test_workflow_template_generates_penalty_and_corporate_rows(client) -> None:  # type: ignore[no-untyped-def]
    template = client.get("/api/bulk/workflow-template")
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.content))
    sheet = workbook.active
    assert sheet is not None
    headers = [cell.value for cell in sheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    sheet.cell(2, columns["Workflow ID"], "SYNTH-PENA-WF-A")
    sheet.cell(2, columns["Message Reference"], "PENASTMTA001")
    sheet.cell(2, columns["Penalty Reference"], "PENALTYA001")
    sheet.cell(3, columns["Workflow ID"], "SYNTH-CA-WF-A")
    sheet.cell(3, columns["Message Reference"], "CA564SYNTHA01")
    sheet.cell(3, columns["Event Reference"], "CAEVSYNTHA01")
    upload = BytesIO()
    workbook.save(upload)
    response = client.post(
        "/api/bulk/workflow-generate",
        files={
            "file": (
                "synthetic-workflows.xlsx",
                upload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["generatedRows"] == 2
    assert payload["failedRows"] == 0
    assert {row["resolvedMessageType"] for row in payload["rowResults"]} == {
        "MT537",
        "MT564",
    }
    archive_response = client.get(payload["downloadPath"])
    with ZipFile(BytesIO(archive_response.content)) as archive:
        names = set(archive.namelist())
    assert "execution-report.json" in names
    assert "summary.xlsx" in names
    assert any(name.endswith("_MT537.txt") for name in names)
    assert any(name.endswith("_MT564.txt") for name in names)


def test_workflow_bulk_continues_after_invalid_row_and_escapes_formula(client) -> None:  # type: ignore[no-untyped-def]
    template = client.get("/api/bulk/workflow-template").content
    workbook = load_workbook(BytesIO(template))
    sheet = workbook.active
    assert sheet is not None
    headers = [cell.value for cell in sheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}
    sheet.cell(2, columns["Workflow ID"], "SYNTH-PENA-WF2")
    sheet.cell(2, columns["Message Reference"], "PENASTMT0002")
    sheet.cell(2, columns["Penalty Reference"], "PENALTY0002")
    sheet.cell(3, columns["Workflow ID"], "SYNTH-CA-WF2")
    sheet.cell(3, columns["Message Reference"], "CA564SYNTH002")
    sheet.cell(3, columns["Event Reference"], "CAEVSYNTH002")
    row = {header: None for header in headers}
    row.update(
        {
            "Workflow Type": "UNSUPPORTED",
            "Scenario ID": "=2+2",
            "Profile ID": "BASE_DEMO_V1",
            "Workflow ID": "BAD-WF",
            "Message Reference": "BADREF0001",
            "Related Message Reference": "NONE",
        }
    )
    sheet.append([row[header] for header in headers])
    content = BytesIO()
    workbook.save(content)
    response = client.post(
        "/api/bulk/workflow-generate",
        files={
            "file": (
                "mixed-workflows.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["generatedRows"] == 2
    assert payload["failedRows"] == 1
    assert payload["rowResults"][-1]["status"] == "FAILED"


def test_workflow_bulk_rejects_invalid_file(client) -> None:  # type: ignore[no-untyped-def]
    response = client.post(
        "/api/bulk/workflow-generate",
        files={
            "file": (
                "not-workbook.xlsx",
                b"not a zip",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 400


def test_workflow_bulk_generates_correlated_corporate_action_rows(client) -> None:  # type: ignore[no-untyped-def]
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(HEADERS)
    base = {
        "Scenario ID": "CA-BULK",
        "Profile ID": "BASE_DEMO_V1",
        "Workflow ID": "CA-BULK-WF",
    }
    rows = [
        base
        | {
            "Workflow Type": "CORPORATE_NOTIFICATION",
            "Message Reference": "CA564BULK001",
            "Related Message Reference": "NONE",
            "Event Reference": "CAEVEVBULK001",
            "Security": "XS0000000001",
            "Safekeeping Account": "SYNTHSAFE01",
            "Eligible Quantity": "1000",
            "Election Deadline": "2099-08-10",
            "Payment Date": "2099-08-15",
        },
        base
        | {
            "Workflow Type": "CORPORATE_INSTRUCTION",
            "Message Reference": "CA565BULK001",
            "Related Message Reference": "CA564BULK001",
            "Option Number": 1,
            "Quantity": "800",
        },
        base
        | {
            "Workflow Type": "CORPORATE_STATUS",
            "Message Reference": "CA567BULK001",
            "Related Message Reference": "CA565BULK001",
            "Status": "PENDING",
        },
        base
        | {
            "Workflow Type": "CORPORATE_CONFIRMATION",
            "Message Reference": "CA566BULK001",
            "Related Message Reference": "CA565BULK001",
            "Option Number": 1,
            "Quantity": "800",
            "Currency": "USD",
            "Amount": "125.50",
            "Payment Date": "2099-08-15",
        },
        base
        | {
            "Workflow Type": "CORPORATE_NARRATIVE",
            "Message Reference": "CA568BULK001",
            "Related Message Reference": "CA564BULK001",
            "Narrative": "SYNTHETIC BULK NARRATIVE.",
        },
    ]
    for row in rows:
        sheet.append([row.get(header) for header in HEADERS])
    content = BytesIO()
    workbook.save(content)
    response = client.post(
        "/api/bulk/workflow-generate",
        files={
            "file": (
                "corporate-lifecycle.xlsx",
                content.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["generatedRows"] == 5
    assert payload["failedRows"] == 0
    assert [row["resolvedMessageType"] for row in payload["rowResults"]] == [
        "MT564",
        "MT565",
        "MT567",
        "MT566",
        "MT568",
    ]
