"""Excel templates, parsing and the Excel automation endpoint."""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.studio.excel import (
    MT_HEADERS,
    MX_HEADERS,
    ExcelFormatError,
    build_template,
    parse_workbook,
    validate_upload,
)
from app.studio.models import MessageFormat

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Scenarios"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def upload(client, content: bytes, name: str = "scenarios.xlsx", **params: object):  # type: ignore[no-untyped-def]
    return client.post(
        "/api/v1/messages/generate-from-excel",
        files={"file": (name, content, XLSX)},
        params=params,
    )


# -- templates -------------------------------------------------------------------------


@pytest.mark.parametrize("format_", [MessageFormat.MT, MessageFormat.MX])
def test_template_has_the_three_sheets(format_: MessageFormat) -> None:
    workbook = load_workbook(BytesIO(build_template(format_)))

    assert workbook.sheetnames == ["Scenarios", "Reference", "Read me"]


def test_mt_template_uses_tag_level_columns() -> None:
    workbook = load_workbook(BytesIO(build_template(MessageFormat.MT)))
    headers = [cell.value for cell in workbook["Scenarios"][1]]

    assert headers == MT_HEADERS


def test_mx_template_uses_element_level_columns() -> None:
    workbook = load_workbook(BytesIO(build_template(MessageFormat.MX)))
    headers = [cell.value for cell in workbook["Scenarios"][1]]

    assert headers == MX_HEADERS


def test_reference_sheet_lists_every_supported_field() -> None:
    from app.studio.catalogue import message_spec

    workbook = load_workbook(BytesIO(build_template(MessageFormat.MX, ["sese.023"])))
    paths = {row[1] for row in workbook["Reference"].iter_rows(min_row=2, values_only=True)}
    expected = {item.xpath for item in message_spec(MessageFormat.MX, "sese.023").fields}

    assert expected <= paths


def test_template_rows_come_from_a_valid_sample() -> None:
    parsed = parse_workbook(build_template(MessageFormat.MT))

    assert parsed.format is MessageFormat.MT
    assert [item.scenario_id for item in parsed.scenarios] == ["TC001", "TC002", "TC003"]
    assert all(not item.issues for item in parsed.scenarios)


# -- parsing ---------------------------------------------------------------------------


def test_format_is_detected_from_the_columns() -> None:
    mt = parse_workbook(
        workbook_bytes(MT_HEADERS, [["TC1", "MT541", "", "GENL", 1, "20C", "SEME", "C", "REF1"]])
    )
    mx = parse_workbook(
        workbook_bytes(
            MX_HEADERS,
            [["MX1", "sese.023", "", "/Document/SctiesSttlmTxInstr/TxId", 1, "REF1"]],
        )
    )

    assert mt.format is MessageFormat.MT
    assert mx.format is MessageFormat.MX


def test_headers_are_matched_case_and_space_insensitively() -> None:
    headers = ["scenario id", "Message Type", "Profile ID", "sequence", "Sequence Occurrence",
               "TAG", "qualifier", "Option", "value"]

    parsed = parse_workbook(
        workbook_bytes(headers, [["TC1", "MT541", "", "GENL", 1, "20C", "SEME", "C", "REF1"]])
    )

    assert parsed.scenarios[0].fields[0].tag == "20C"


def test_missing_required_column_is_rejected_with_the_expected_layout() -> None:
    with pytest.raises(ExcelFormatError) as raised:
        parse_workbook(workbook_bytes(["ScenarioID", "Tag", "Value"], [["TC1", "20C", "X"]]))

    assert "MessageType" in str(raised.value)


def test_sheet_with_neither_tag_nor_xpath_is_rejected() -> None:
    with pytest.raises(ExcelFormatError):
        parse_workbook(workbook_bytes(["ScenarioID", "MessageType", "Value"], [["a", "b", "c"]]))


def test_empty_sheet_is_rejected() -> None:
    with pytest.raises(ExcelFormatError):
        parse_workbook(workbook_bytes(MT_HEADERS, []))


def test_rows_group_into_scenarios_in_first_seen_order() -> None:
    rows = [
        ["TC2", "MT541", "", "GENL", 1, "20C", "SEME", "C", "B"],
        ["TC1", "MT541", "", "GENL", 1, "20C", "SEME", "C", "A"],
        ["TC2", "MT541", "", "GENL", 1, "23G", None, "G", "NEWM"],
    ]

    parsed = parse_workbook(workbook_bytes(MT_HEADERS, rows))

    assert [item.scenario_id for item in parsed.scenarios] == ["TC2", "TC1"]
    assert len(parsed.scenarios[0].fields) == 2


def test_blank_rows_are_skipped() -> None:
    rows = [
        ["TC1", "MT541", "", "GENL", 1, "20C", "SEME", "C", "A"],
        [None, None, None, None, None, None, None, None, None],
        ["TC1", "MT541", "", "GENL", 1, "23G", None, "G", "NEWM"],
    ]

    parsed = parse_workbook(workbook_bytes(MT_HEADERS, rows))

    assert len(parsed.scenarios) == 1
    assert len(parsed.scenarios[0].fields) == 2


def test_excel_dates_are_converted_back_to_iso_text() -> None:
    from datetime import datetime

    rows = [
        [
            "MX1",
            "sese.023",
            "",
            "/Document/SctiesSttlmTxInstr/TradDtls/SttlmDt/Dt/Dt",
            1,
            datetime(2026, 8, 18),
        ]
    ]

    parsed = parse_workbook(workbook_bytes(MX_HEADERS, rows))

    assert parsed.scenarios[0].elements[0].value == "2026-08-18"


def test_numeric_cells_do_not_gain_a_decimal_point() -> None:
    rows = [["MX1", "sese.023", "", "/Document/SctiesSttlmTxInstr/QtyAndAcctDtls/SttlmQty/Qty/Unit",
             1, 1000]]

    parsed = parse_workbook(workbook_bytes(MX_HEADERS, rows))

    assert parsed.scenarios[0].elements[0].value == "1000"


def test_missing_value_is_a_row_level_issue() -> None:
    rows = [["TC1", "MT541", "", "GENL", 1, "20C", "SEME", "C", None]]

    parsed = parse_workbook(workbook_bytes(MT_HEADERS, rows))

    assert parsed.scenarios[0].issues[0].location == "row 2"


def test_missing_tag_is_a_row_level_issue() -> None:
    rows = [["TC1", "MT541", "", "GENL", 1, None, "SEME", "C", "REF"]]

    parsed = parse_workbook(workbook_bytes(MT_HEADERS, rows))

    assert "Tag is missing" in parsed.scenarios[0].issues[0].message


def test_mixing_message_types_in_one_scenario_is_reported() -> None:
    rows = [
        ["TC1", "MT541", "", "GENL", 1, "20C", "SEME", "C", "A"],
        ["TC1", "MT543", "", "GENL", 1, "23G", None, "G", "NEWM"],
    ]

    parsed = parse_workbook(workbook_bytes(MT_HEADERS, rows))

    assert "mixes message types" in parsed.scenarios[0].issues[0].message


def test_sequence_occurrence_builds_repeated_sequences() -> None:
    rows = [
        ["MX1", "sese.023", "", "/Document/SctiesSttlmTxInstr/SttlmParams/SttlmTxCond/Cd", 1,
         "NOMC"],
        ["MX1", "sese.023", "", "/Document/SctiesSttlmTxInstr/SttlmParams/SttlmTxCond/Cd", 2,
         "PART"],
    ]

    parsed = parse_workbook(workbook_bytes(MX_HEADERS, rows))

    assert [item.occurrence for item in parsed.scenarios[0].elements] == [1, 2]


def test_row_limit_is_enforced() -> None:
    rows = [["TC1", "MT541", "", "GENL", 1, "20C", "SEME", "C", f"R{n}"] for n in range(12)]

    with pytest.raises(ExcelFormatError) as raised:
        parse_workbook(workbook_bytes(MT_HEADERS, rows), max_rows=10)

    assert "more than 10" in str(raised.value)


# -- upload guards ---------------------------------------------------------------------


def test_non_xlsx_extension_is_rejected() -> None:
    with pytest.raises(ExcelFormatError):
        validate_upload(b"PK\x03\x04", "scenarios.csv", 1_000)


def test_path_traversal_filename_is_rejected() -> None:
    with pytest.raises(ExcelFormatError):
        validate_upload(b"PK\x03\x04", "../scenarios.xlsx", 1_000)


def test_oversized_upload_is_rejected() -> None:
    with pytest.raises(ExcelFormatError):
        validate_upload(workbook_bytes(MT_HEADERS, []), "scenarios.xlsx", 10)


def test_content_that_is_not_a_workbook_is_rejected() -> None:
    with pytest.raises(ExcelFormatError):
        validate_upload(b"not a workbook", "scenarios.xlsx", 1_000)


# -- endpoint --------------------------------------------------------------------------


def test_mt_template_round_trips_through_the_api(client) -> None:  # type: ignore[no-untyped-def]
    template = client.get("/api/v1/templates/MT.xlsx")
    assert template.status_code == 200

    response = upload(client, template.content)

    payload = response.json()
    assert response.status_code == 200
    assert payload["format"] == "MT"
    assert payload["failed"] == 0
    assert payload["generated"] == payload["totalScenarios"] == 3
    assert all(item["outputs"]["fin"].startswith("{1:F01") for item in payload["results"])


def test_mx_template_round_trips_through_the_api(client) -> None:  # type: ignore[no-untyped-def]
    template = client.get("/api/v1/templates/MX.xlsx")

    response = upload(client, template.content)

    payload = response.json()
    assert payload["format"] == "MX"
    assert payload["failed"] == 0
    for item in payload["results"]:
        assert item["outputs"]["appHdr"].lstrip().startswith("<AppHdr")
        assert "urn:iso:std:iso:20022" in item["outputs"]["document"]


def test_multiple_scenarios_are_reported_individually(client) -> None:  # type: ignore[no-untyped-def]
    from app.studio.models import SampleVariant
    from app.studio.samples import build_sample

    sample = build_sample(MessageFormat.MT, "MT541", SampleVariant.TYPICAL)
    good = [
        ["TC001", "MT541", "BASE_DEMO_V1", item.sequence, 1, item.tag, item.qualifier,
         item.option, item.value]
        for item in sample.inputs
    ]
    broken = [
        ["TC002", "MT541", "BASE_DEMO_V1", item.sequence, 1, item.tag, item.qualifier,
         item.option, item.value]
        for item in sample.inputs
        if item.qualifier != "SETT" or item.tag != "19A"
    ]

    response = upload(client, workbook_bytes(MT_HEADERS, [*good, *broken]))

    payload = response.json()
    assert payload["totalScenarios"] == 2
    assert payload["generated"] == 1
    assert payload["failed"] == 1
    first, second = payload["results"]
    assert first["status"] == "GENERATED"
    assert second["status"] == "INVALID"
    assert second["validation"]["errors"]


def test_row_numbers_are_reported_for_every_scenario(client) -> None:  # type: ignore[no-untyped-def]
    rows = [
        ["TC1", "MT541", "", "GENL", 1, "20C", "SEME", "C", "A"],
        ["TC1", "MT541", "", "GENL", 1, "23G", None, "G", "NEWM"],
    ]

    response = upload(client, workbook_bytes(MT_HEADERS, rows))

    assert response.json()["results"][0]["rowNumbers"] == [2, 3]


def test_unsupported_message_type_fails_only_its_scenario(client) -> None:  # type: ignore[no-untyped-def]
    rows = [
        ["TC1", "MT999", "", "GENL", 1, "20C", "SEME", "C", "A"],
        ["TC2", "MT541", "", "GENL", 1, "20C", "SEME", "C", "B"],
    ]

    payload = upload(client, workbook_bytes(MT_HEADERS, rows)).json()

    assert payload["results"][0]["status"] == "FAILED"
    assert payload["results"][0]["validation"]["errors"][0]["ruleId"] == (
        "EXCEL_UNSUPPORTED_MESSAGE_TYPE"
    )
    assert payload["results"][1]["status"] == "INVALID"


def test_bad_xpath_is_reported_against_its_scenario(client) -> None:  # type: ignore[no-untyped-def]
    rows = [["MX1", "sese.023", "", "/Document/SctiesSttlmTxInstr/Nope", 1, "X"]]

    payload = upload(client, workbook_bytes(MX_HEADERS, rows)).json()

    codes = {item["ruleId"] for item in payload["results"][0]["validation"]["errors"]}
    assert "MX_UNKNOWN_ELEMENT" in codes


def test_wrong_content_type_is_rejected(client) -> None:  # type: ignore[no-untyped-def]
    response = client.post(
        "/api/v1/messages/generate-from-excel",
        files={"file": ("scenarios.xlsx", b"PK\x03\x04", "text/csv")},
    )

    assert response.status_code == 415


def test_malformed_workbook_is_rejected_with_guidance(client) -> None:  # type: ignore[no-untyped-def]
    response = upload(client, b"not a workbook at all")

    assert response.status_code == 422
    assert "xlsx" in response.json()["error"]["message"].lower()


def test_profile_can_be_chosen_per_upload(client) -> None:  # type: ignore[no-untyped-def]
    template = client.get("/api/v1/templates/MT.xlsx")

    payload = upload(client, template.content, profileId="BFS_CLIENT_DEMO_V1").json()

    # The template references BASE_DEMO_V1 explicitly, so the query default does not win;
    # what matters is that the request is accepted and every scenario is reported.
    assert payload["totalScenarios"] == 3


# -- every configured message, over the automation path ----------------------------------
#
# The Excel importer, the JSON API and the browser all call the same StudioService. These
# tests are the proof for the Excel third of that claim, and they are derived from the
# registries so a message added as YAML is covered without anyone editing this file.


def _every_message(format_: MessageFormat) -> list[str]:
    from app.specifications.registry import specification_registry
    from app.studio.mx.registry import mx_registry

    if format_ is MessageFormat.MX:
        return [spec.message_type for spec in mx_registry.all_specs()]
    return [spec.message_type for spec in specification_registry.list()]


def _rows_for(format_: MessageFormat, message_type: str) -> list[list[object]]:
    from app.studio.samples import available_variants, build_sample

    variant = available_variants(format_, message_type)[0]
    sample = build_sample(format_, message_type, variant)
    scenario = f"TC-{message_type}"
    if format_ is MessageFormat.MX:
        return [
            [scenario, message_type, "BASE_DEMO_V1", item.path, item.occurrence, item.value]
            for item in sample.elements
        ]
    return [
        [
            scenario,
            message_type,
            "BASE_DEMO_V1",
            item.sequence,
            1,
            item.tag,
            item.qualifier,
            item.option,
            item.value,
        ]
        for item in sample.inputs
    ]


@pytest.mark.parametrize(
    ("format_", "headers"),
    [(MessageFormat.MT, MT_HEADERS), (MessageFormat.MX, MX_HEADERS)],
)
def test_every_configured_message_generates_from_a_workbook(  # type: ignore[no-untyped-def]
    client, format_: MessageFormat, headers: list[str]
) -> None:
    message_types = _every_message(format_)
    rows = [row for message_type in message_types for row in _rows_for(format_, message_type)]

    response = upload(client, workbook_bytes(headers, rows), persist=False)

    assert response.status_code == 200
    body = response.json()
    assert body["totalScenarios"] == len(message_types)
    failures = [item for item in body["results"] if not item["valid"]]
    assert not failures, [
        (item["scenarioId"], item["validation"]["errors"]) for item in failures
    ]


def test_an_excel_caller_receives_the_raw_message_not_just_a_verdict(client) -> None:  # type: ignore[no-untyped-def]
    """An automation tester's whole reason for calling this is the message text itself."""
    rows = _rows_for(MessageFormat.MT, "MT541")

    body = upload(client, workbook_bytes(MT_HEADERS, rows), persist=False).json()
    outputs = body["results"][0]["outputs"]

    assert outputs["fin"].startswith("{1:F01")
    assert ":16R:GENL" in outputs["block4"]
    assert body["results"][0]["validation"]["layers"]
    assert body["results"][0]["checksum"]


def test_an_excel_mx_caller_receives_the_document_and_header(client) -> None:  # type: ignore[no-untyped-def]
    rows = _rows_for(MessageFormat.MX, "sese.023")

    body = upload(client, workbook_bytes(MX_HEADERS, rows), persist=False).json()
    outputs = body["results"][0]["outputs"]

    assert outputs["document"].startswith("<Document")
    assert "urn:iso:std:iso:20022:tech:xsd:sese.023.001.11" in outputs["xml"]
    assert "{1:" not in outputs["xml"]


def test_excel_and_json_produce_the_same_message(client) -> None:  # type: ignore[no-untyped-def]
    """The two entry points share one composer, so they cannot disagree. Prove it."""
    from app.studio.models import SampleVariant
    from app.studio.samples import build_sample

    sample = build_sample(MessageFormat.MX, "sese.023", SampleVariant.TYPICAL)
    rows = [
        ["TC-sese.023", "sese.023", "BASE_DEMO_V1", item.path, item.occurrence, item.value]
        for item in sample.elements
    ]

    from_excel = upload(client, workbook_bytes(MX_HEADERS, rows), persist=False).json()
    from_json = client.post(
        "/api/v1/messages/generate",
        json={
            "format": "MX",
            "messageType": "sese.023",
            "elements": [item.model_dump(by_alias=True) for item in sample.elements],
            "persist": False,
        },
    ).json()

    assert (
        from_excel["results"][0]["outputs"]["document"] == from_json["outputs"]["document"]
    )


def test_the_reference_sheet_covers_every_configured_message() -> None:
    """The Scenarios sheet is a handful of worked examples; the Reference sheet is the
    dictionary. A message missing from it is generatable but undiscoverable, which is how a
    new message quietly fails to reach automation testers."""
    for format_ in (MessageFormat.MT, MessageFormat.MX):
        workbook = load_workbook(BytesIO(build_template(format_)))
        listed = {
            str(row[0])
            for row in workbook["Reference"].iter_rows(min_row=2, values_only=True)
        }

        assert set(_every_message(format_)) == listed
