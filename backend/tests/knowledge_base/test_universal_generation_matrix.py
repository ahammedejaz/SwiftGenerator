"""The universal smoke harness: every generation-ready MT structure, through the ordinary API.

No message is named here. Whatever the synced knowledge base marks ``GENERATION_READY`` is
taken through sample → generate (FIN) → import → round trip → Excel template → JSON, the
same endpoints the browser and automation call. In CI the knowledge base is the synthetic
fixture corpus plus the committed Prowide subset; on an operator's machine
(``KNOWLEDGE_MATRIX_DB`` pointing at a real database) it is every structure the real
sources compiled.
"""

from __future__ import annotations

import io
import os
import pathlib
from collections.abc import Iterator

import pytest
from openpyxl import load_workbook

from tests.knowledge_fixtures import knowledge_client, knowledge_env  # noqa: F401 - fixtures

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _ready_entries(client) -> list[dict]:  # type: ignore[no-untyped-def]
    catalogue = client.get("/api/v1/catalogue").json()
    return [
        entry
        for entry in catalogue["messages"]
        if entry["format"] == "MT" and entry["lane"] == "KNOWLEDGE_PREVIEW" and entry["generatable"]
    ]


def _matrix(client, entry: dict) -> None:  # type: ignore[no-untyped-def]
    message_type, release = entry["messageType"], entry["release"]
    params = {"format": "MT", "lane": "KNOWLEDGE_PREVIEW", "release": release}
    sample = client.get(f"/api/v1/messages/{message_type}/samples/MINIMAL", params=params)
    assert sample.status_code == 200, (message_type, release, sample.text)
    inputs = sample.json()["inputs"]
    generated = client.post(
        "/api/v1/messages/generate",
        json={
            "format": "MT",
            "messageType": message_type,
            "fields": inputs,
            "persist": False,
            "lane": "KNOWLEDGE_PREVIEW",
            "release": release,
            "outputModes": ["BLOCK4", "FIN"],
        },
    )
    assert generated.status_code == 200, (message_type, release, generated.text)
    result = generated.json()
    assert result["valid"], (message_type, release, result["validation"])
    block4 = result["outputs"]["block4"]
    assert block4.startswith("{4:") and block4.rstrip().endswith("-}")
    assert result["outputs"].get("fin"), "FIN envelope from profile data"
    # JSON in, JSON out: the canonical values are what the response echoes.
    imported = client.post(
        "/api/v1/messages/import",
        json={
            "text": result["outputs"]["fin"],
            "messageType": message_type,
            "lane": "KNOWLEDGE_PREVIEW",
            "release": release,
            "persist": False,
        },
    )
    assert imported.status_code == 200, (message_type, release, imported.text)
    body = imported.json()
    assert body["importIssues"] == [], (message_type, release, body["importIssues"])
    assert body["result"]["outputs"]["block4"] == block4, (message_type, release)
    # Excel: the template for this structure round-trips the same values.
    template = client.get(
        "/api/v1/templates/MT.xlsx",
        params={"messageType": message_type, "lane": "KNOWLEDGE_PREVIEW", "release": release},
    )
    assert template.status_code == 200, (message_type, release, template.text)
    workbook = load_workbook(io.BytesIO(template.content))
    assert "Scenarios" in workbook.sheetnames
    uploaded = client.post(
        "/api/v1/messages/generate-from-excel",
        params={"lane": "KNOWLEDGE_PREVIEW", "release": release, "persist": "false"},
        files={"file": ("t.xlsx", template.content, XLSX)},
    )
    assert uploaded.status_code == 200, (message_type, release, uploaded.text)
    excel = uploaded.json()
    assert excel["generated"] == excel["totalScenarios"] >= 1, (message_type, release)


def test_every_generation_ready_structure_passes_the_matrix(knowledge_client) -> None:  # type: ignore[no-untyped-def]  # noqa: F811
    entries = _ready_entries(knowledge_client)
    assert entries, "the fixture corpus compiles generation-ready structures"
    failures: list[str] = []
    for entry in entries:
        try:
            _matrix(knowledge_client, entry)
        except AssertionError as error:  # noqa: PERF203 - collect every failure, then report
            failures.append(f"{entry['messageType']}/{entry['release']}: {str(error)[:300]}")
    assert not failures, "\n".join(failures)


def test_no_blocked_structure_generates(knowledge_client) -> None:  # type: ignore[no-untyped-def]  # noqa: F811
    """A structure the gates refused is refused by the API too, with the same code."""
    catalogue = knowledge_client.get("/api/v1/catalogue").json()
    blocked = [
        entry
        for entry in catalogue["messages"]
        if entry["format"] == "MT"
        and entry["lane"] == "KNOWLEDGE_PREVIEW"
        and not entry["generatable"]
    ]
    for entry in blocked[:10]:
        response = knowledge_client.post(
            "/api/v1/messages/generate",
            json={
                "format": "MT",
                "messageType": entry["messageType"],
                "release": entry["release"],
                "fields": [],
                "persist": False,
                "lane": "KNOWLEDGE_PREVIEW",
            },
        )
        assert response.status_code in {404, 409, 422}, (entry["messageType"], response.text)


@pytest.fixture
def real_matrix_client() -> Iterator[object]:
    """The operator's real knowledge database, when ``KNOWLEDGE_MATRIX_DB`` names one."""
    path = os.environ.get("KNOWLEDGE_MATRIX_DB")
    if not path:
        pytest.skip("KNOWLEDGE_MATRIX_DB not set: the real-corpus matrix runs locally only")
    from fastapi.testclient import TestClient

    from app.main import app
    from tests.knowledge_fixtures import _rewire

    keys = ("KNOWLEDGE_MODE", "KNOWLEDGE_DB_PATH", "KNOWLEDGE_PACK_DIR", "KNOWLEDGE_AI_PROVIDER")
    previous = {key: os.environ.get(key) for key in keys}
    pack_dir = os.environ.get("KNOWLEDGE_MATRIX_PACKS", str(pathlib.Path(path).parent / "packs"))
    _rewire(
        {
            "KNOWLEDGE_MODE": "local",
            "KNOWLEDGE_DB_PATH": path,
            "KNOWLEDGE_PACK_DIR": pack_dir,
            "KNOWLEDGE_AI_PROVIDER": "scripted",
        }
    )
    with TestClient(app) as client:
        yield client
    restore: dict[str, str] = {}
    for key, value in previous.items():
        if value is None:
            os.environ.pop(key, None)
        else:
            restore[key] = value
    _rewire(restore)


def test_the_real_corpus_matrix(real_matrix_client) -> None:  # type: ignore[no-untyped-def]
    entries = _ready_entries(real_matrix_client)
    failures: list[str] = []
    for entry in entries:
        try:
            _matrix(real_matrix_client, entry)
        except AssertionError as error:  # noqa: PERF203
            failures.append(f"{entry['messageType']}/{entry['release']}: {str(error)[:300]}")
    print(f"real corpus matrix: {len(entries) - len(failures)}/{len(entries)} passed")
    assert not failures, "\n".join(failures)
