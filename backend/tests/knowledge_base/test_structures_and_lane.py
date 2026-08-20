"""Dynamic Structure Packs and the KNOWLEDGE_PREVIEW lane, end to end.

The critical proofs: a message that was never configured in this repository — MT999 from a
synthetic guide, MT103 from Prowide evidence, test.001 from a synthetic XSD — becomes
generation-ready through the ordinary StudioService, forms, Excel, JSON API and import,
with no message-specific Python. The lane is explicit everywhere and the configured lane is
untouched.
"""

from __future__ import annotations

import re
from pathlib import Path

import pytest

from app.knowledge_base.structures.swift_format import (
    FormatUnsupported,
    compile_format,
    input_kind_for,
    is_value_less,
    synthetic_value,
)

PACKAGE = Path(__file__).resolve().parents[2] / "app" / "knowledge_base"


# -- the format compiler ---------------------------------------------------------------------


@pytest.mark.parametrize(
    ("notation", "accepted", "rejected"),
    [
        (":4!c//16x", ["SYNTHREF000001"], ["", "X" * 17, "bad char ü"]),
        (":4!c//<DATE4>", ["20260818"], ["2026-08-18", "2026081"]),
        (":4!c//[<N>]<CUR><AMOUNT>15", ["USD1000,", "NUSD1,5"], ["USD1000", "USD1,000.00"]),
        (":4!c//4!c/<AMOUNT>15", ["UNIT/100,"], ["UNIT100,", "UNIT/100"]),
        (":4!c/8c/34x", ["SYNTH/ID01"], ["ID01"]),
        (":4!c//<BIC>", ["SYNTGB2L", "SYNTGB2LXXX"], ["SYNT", "SYNTGB2LXXXX1"]),
        ("4!c[/4!c]", ["NEWM", "NEWM/CODU"], ["NEW", "NEWM/CO"]),
        ("<DATE2><CUR><AMOUNT>15", ["260818USD500,"], ["20260818USD500,"]),
        ("[/34x$]35x[$35x]0-3", ["/ACC\nNAME", "NAME ONLY"], ["A" * 36]),
        ("6!n", ["123456"], ["12345", "1234567"]),
        ("<DC><DATE2><3!a><AMOUNT>15", ["C260818USD1,"], ["X260818USD1,"]),
    ],
)
def test_swift_notation_compiles_to_a_faithful_pattern(
    notation: str, accepted: list[str], rejected: list[str]
) -> None:
    compiled = compile_format(notation)
    for value in accepted:
        assert re.fullmatch(compiled.pattern, value), (notation, value)
    for value in rejected:
        assert not re.fullmatch(compiled.pattern, value), (notation, value)


def test_qualifier_separator_and_optional_dss_follow_the_notation() -> None:
    assert compile_format(":4!c//16x").qualifier_separator == "//"
    assert compile_format(":4!c/8c/34x").qualifier_separator == "/"
    optional = compile_format(":4!c/[8c]/4!c")
    assert optional.qualifier_separator == "//" and optional.optional_dss_dropped


def test_unknown_notation_is_refused_not_guessed() -> None:
    with pytest.raises(FormatUnsupported):
        compile_format("<VAR-SEQU-4>")
    with pytest.raises(FormatUnsupported):
        compile_format("[3!c]*10")


def test_synthetic_values_satisfy_their_own_pattern_and_never_claim_reality() -> None:
    for notation in (":4!c//16x", ":4!c//<BIC>", "[<ISIN> 12!c][$][35x][$35x]0-3", ":4!c//<DATE4>"):
        value = synthetic_value(notation)
        assert re.fullmatch(compile_format(notation).pattern, value), notation
    assert synthetic_value(":4!c//<BIC>").startswith("SYNT")
    assert synthetic_value("4!c[/4!c]", codes=["NEWM", "CANC"]) == "NEWM"
    assert is_value_less("$") and is_value_less("") and not is_value_less("16x")


def test_input_kinds_come_from_the_format_not_a_model() -> None:
    assert input_kind_for(compile_format(":4!c//<DATE4>"), codes=False) == "DATE"
    assert input_kind_for(compile_format(":4!c//<BIC>"), codes=False) == "PARTY_BIC"
    assert input_kind_for(compile_format(":4!c//[<N>]<CUR><AMOUNT>15"), codes=False) == "AMOUNT"
    assert input_kind_for(compile_format(":4!c/[8c]/4!c"), codes=True) == "SELECT"


def test_the_compiler_package_names_no_message_type() -> None:
    """No `if MT103`, no `if "pacs.008"`: the only message literals are documentation."""
    offenders: list[str] = []
    for path in (PACKAGE / "structures").glob("*.py"):
        for number, line in enumerate(path.read_text().splitlines(), start=1):
            code = line.split("#", 1)[0]
            if re.search(r"""(==|!=|in)\s*["'](MT\d{3}|[a-z]{4}\.\d{3})["']""", code):
                offenders.append(f"{path.name}:{number}")
    assert offenders == []


# -- the preview lane through the ordinary engine --------------------------------------------


def test_the_catalogue_lists_configured_preview_and_knowledge_only_entries(
    knowledge_client,
) -> None:  # type: ignore[no-untyped-def]
    entries = knowledge_client.get("/api/v1/catalogue").json()["messages"]
    by_key = {(e["format"], e["messageType"], e["lane"], e["release"]): e for e in entries}
    configured = by_key[("MT", "MT541", "CONFIGURED", None)]
    assert configured["generatable"] and configured["readinessLabel"] == "Configured & validated"
    mt999 = by_key[("MT", "MT999", "KNOWLEDGE_PREVIEW", "SR2026")]
    assert mt999["generatable"] and mt999["readiness"] == "GENERATION_READY"
    assert "future release" in mt999["readinessLabel"]
    assert mt999["releaseLane"] == "FUTURE_TEST"
    assert by_key[("MT", "MT999", "KNOWLEDGE_PREVIEW", "SR2027")]["generatable"]
    mt103 = by_key[("MT", "MT103", "KNOWLEDGE_PREVIEW", "SR2025")]
    assert mt103["generatable"] and mt103["structureSource"] == "PROWIDE_SR2025"
    mx = by_key[("MX", "test.001", "KNOWLEDGE_PREVIEW", "test.001.001.01")]
    assert mx["generatable"] and mx["structureSource"] == "OPERATOR_SUPPLIED_XSD"
    knowledge_only = next(e for e in entries if e["messageType"] == "MT998")
    assert knowledge_only["lane"] == "KNOWLEDGE_PREVIEW" and not knowledge_only["generatable"]
    assert knowledge_only["blockers"] and knowledge_only["knowledgeSources"] >= 1
    assert "not" in knowledge_only["readinessLabel"].lower()
    # The Prowide-only MT541 SR2025 structure is shadowed by the configured MT541: the
    # configured pack is the authority for the current live release, so the catalogue
    # does not list a second, weaker MT541 beside it. The status endpoint still tells.
    assert ("MT", "MT541", "KNOWLEDGE_PREVIEW", "SR2025") not in by_key
    status = knowledge_client.get("/api/v1/knowledge/messages/MT541/status").json()
    sr2025 = next(e for e in status["entries"] if e["release"] == "SR2025")
    assert sr2025["readiness"] == "STRUCTURE_AVAILABLE"
    assert "QUALIFIER_EVIDENCE_MISSING" in sr2025["blockers"]
    # Configured entries are exactly the 23 that existed before; nothing was promoted.
    assert sum(1 for e in entries if e["lane"] == "CONFIGURED") == 23
    formats = {f["id"]: f for f in knowledge_client.get("/api/v1/catalogue").json()["formats"]}
    assert (
        formats["MT"]["configuredMessageCount"] == 16
        and formats["MX"]["configuredMessageCount"] == 7
    )


def _sample_and_generate(
    client, format_: str, message_type: str, release: str | None, variant: str = "MINIMAL"
):  # type: ignore[no-untyped-def]
    params = {"format": format_, "lane": "KNOWLEDGE_PREVIEW"}
    if release:
        params["release"] = release
    sample = client.get(f"/api/v1/messages/{message_type}/samples/{variant}", params=params)
    assert sample.status_code == 200, sample.text
    body = sample.json()
    generated = client.post(
        "/api/v1/messages/generate",
        json={
            "format": format_,
            "messageType": message_type,
            "fields": body["inputs"],
            "elements": body["elements"],
            "persist": False,
            "lane": "KNOWLEDGE_PREVIEW",
            "release": release,
        },
    )
    assert generated.status_code == 200, generated.text
    return body, generated.json()


def test_new_mt_from_a_guide_without_code_generates_imports_and_round_trips(
    knowledge_client,
) -> None:  # type: ignore[no-untyped-def]
    spec = knowledge_client.get(
        "/api/v1/messages/MT999/spec", params={"lane": "KNOWLEDGE_PREVIEW", "release": "SR2026"}
    ).json()
    assert spec["lane"] == "KNOWLEDGE_PREVIEW" and spec["release"] == "SR2026"
    assert any(f["tag"] == "95P" and f["qualifier"] == "PSET" for f in spec["fields"])
    assert spec["capabilityStatement"].startswith("Structure-backed test generation")
    sample, result = _sample_and_generate(knowledge_client, "MT", "MT999", "SR2026")
    assert result["valid"], result["validation"]["errors"]
    assert result["lane"] == "KNOWLEDGE_PREVIEW"
    assert result["provenance"]["ruleStatus"] == "NOT_ESTABLISHED"
    assert result["provenance"]["release"] == "SR2026"
    fin = result["outputs"]["fin"]
    assert "{2:I999" in fin and ":20C::SEME//" in fin
    imported = knowledge_client.post(
        "/api/v1/messages/import",
        json={
            "text": fin,
            "messageType": "MT999",
            "lane": "KNOWLEDGE_PREVIEW",
            "release": "SR2026",
            "persist": False,
        },
    )
    assert imported.status_code == 200, imported.text
    body = imported.json()
    assert body["result"]["outputs"]["block4"] == result["outputs"]["block4"]
    assert body["importIssues"] == []


def test_new_mt_from_prowide_evidence_without_code(knowledge_client) -> None:  # type: ignore[no-untyped-def]
    """MT103 has no ``:16R:`` markers at all — the unbracketed body the composer learned."""
    sample, result = _sample_and_generate(knowledge_client, "MT", "MT103", "SR2025")
    assert result["valid"], result["validation"]["errors"]
    fin = result["outputs"]["fin"]
    assert "{2:I103" in fin and ":16R:" not in fin
    assert ":32A:" in fin and ":20:" in fin
    imported = knowledge_client.post(
        "/api/v1/messages/import",
        json={
            "text": fin,
            "messageType": "MT103",
            "lane": "KNOWLEDGE_PREVIEW",
            "release": "SR2025",
            "persist": False,
        },
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["result"]["outputs"]["block4"] == result["outputs"]["block4"]


def test_unbracketed_sequences_with_implicit_boundaries_round_trip(knowledge_client) -> None:  # type: ignore[no-untyped-def]
    """MT101's Sequence B repeats on its leading tag; MT204's ``20`` opens two sequences."""
    for message_type in ("MT101", "MT204", "MT935"):
        sample, result = _sample_and_generate(
            knowledge_client, "MT", message_type, "SR2025", "FULL"
        )
        assert result["valid"], (message_type, result["validation"]["errors"][:2])
        imported = knowledge_client.post(
            "/api/v1/messages/import",
            json={
                "text": result["outputs"]["block4"],
                "messageType": message_type,
                "lane": "KNOWLEDGE_PREVIEW",
                "release": "SR2025",
                "persist": False,
            },
        )
        assert imported.status_code == 200, imported.text
        assert imported.json()["result"]["outputs"]["block4"] == result["outputs"]["block4"], (
            message_type
        )


def test_value_less_markers_are_written_by_the_composer_never_supplied(knowledge_client) -> None:  # type: ignore[no-untyped-def]
    spec = knowledge_client.get("/api/v1/messages/MT300/spec", params={"lane": "KNOWLEDGE_PREVIEW"})
    if spec.status_code != 200:
        pytest.skip("MT300 is not generation-ready in this evidence")
    assert not any(f["tag"] == "15A" for f in spec.json()["fields"]), (
        "markers are not form controls"
    )
    _sample, result = _sample_and_generate(knowledge_client, "MT", "MT300", "SR2025")
    assert ":15A:\n" in result["outputs"]["block4"] or ":15A:\r\n" in result["outputs"]["block4"]


def test_new_mx_from_an_xsd_without_code(knowledge_client) -> None:  # type: ignore[no-untyped-def]
    sample, result = _sample_and_generate(knowledge_client, "MX", "test.001.001.01", None)
    assert result["valid"], result["validation"]["errors"]
    layers = {layer["layer"]: layer for layer in result["validation"]["layers"]}
    assert layers["XSD"]["state"] == "PASSED"
    assert "operator-supplied" in layers["XSD"]["detail"].lower()
    assert result["provenance"]["structureSource"] == "OPERATOR_SUPPLIED_XSD"
    xml = result["outputs"]["xml"]
    imported = knowledge_client.post(
        "/api/v1/messages/import", json={"text": xml, "lane": "KNOWLEDGE_PREVIEW", "persist": False}
    )
    assert imported.status_code == 200, imported.text
    assert imported.json()["result"]["outputs"]["document"] == result["outputs"]["document"]


def test_excel_template_and_upload_use_the_same_engine_as_json(knowledge_client) -> None:  # type: ignore[no-untyped-def]
    from io import BytesIO

    from openpyxl import load_workbook

    template = knowledge_client.get(
        "/api/v1/templates/MT.xlsx",
        params={"messageType": "MT999", "lane": "KNOWLEDGE_PREVIEW", "release": "SR2026"},
    )
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.content))
    rows = list(workbook["Scenarios"].iter_rows(values_only=True))
    assert rows[0][:9] == (
        "ScenarioID",
        "MessageType",
        "ProfileID",
        "Sequence",
        "SequenceOccurrence",
        "Tag",
        "Qualifier",
        "Option",
        "Value",
    )
    assert any(row[1] == "MT999" for row in rows[1:])
    uploaded = knowledge_client.post(
        "/api/v1/messages/generate-from-excel",
        params={"lane": "KNOWLEDGE_PREVIEW", "release": "SR2026", "persist": "false"},
        files={
            "file": (
                "t.xlsx",
                template.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert uploaded.status_code == 200, uploaded.text
    body = uploaded.json()
    assert body["generated"] == body["totalScenarios"] >= 1
    excel_block4 = body["results"][0]["outputs"]["block4"]
    # The same canonical values through JSON give byte-identical output.
    sample = knowledge_client.get(
        "/api/v1/messages/MT999/samples/TYPICAL",
        params={"format": "MT", "lane": "KNOWLEDGE_PREVIEW", "release": "SR2026"},
    )
    if sample.status_code == 404:
        sample = knowledge_client.get(
            "/api/v1/messages/MT999/samples/MINIMAL",
            params={"format": "MT", "lane": "KNOWLEDGE_PREVIEW", "release": "SR2026"},
        )
    generated = knowledge_client.post(
        "/api/v1/messages/generate",
        json={
            "format": "MT",
            "messageType": "MT999",
            "fields": sample.json()["inputs"],
            "persist": False,
            "lane": "KNOWLEDGE_PREVIEW",
            "release": "SR2026",
        },
    )
    assert generated.json()["outputs"]["block4"] == excel_block4


def test_release_isolation_two_packs_of_one_message_never_merge(knowledge_client) -> None:  # type: ignore[no-untyped-def]
    older = knowledge_client.get(
        "/api/v1/messages/MT999/spec", params={"lane": "KNOWLEDGE_PREVIEW", "release": "SR2026"}
    ).json()
    newer = knowledge_client.get(
        "/api/v1/messages/MT999/spec", params={"lane": "KNOWLEDGE_PREVIEW", "release": "SR2027"}
    ).json()
    older_ids = {f["id"] for f in older["fields"]}
    newer_ids = {f["id"] for f in newer["fields"]}
    assert "MT999-E1-97A-CASH" in newer_ids and "MT999-E1-97A-CASH" not in older_ids
    # Without a release, two generation-ready releases are refused rather than guessed.
    ambiguous = knowledge_client.get(
        "/api/v1/messages/MT999/spec", params={"lane": "KNOWLEDGE_PREVIEW"}
    )
    assert ambiguous.status_code == 404
    assert "KNOWLEDGE_RELEASE_REQUIRED" in ambiguous.json()["error"]["message"]


def test_knowledge_only_messages_block_generation_with_the_exact_reason(knowledge_client) -> None:  # type: ignore[no-untyped-def]
    blocked = knowledge_client.post(
        "/api/v1/messages/generate",
        json={
            "format": "MT",
            "messageType": "MT541",
            "release": "SR2025",
            "fields": [],
            "persist": False,
            "lane": "KNOWLEDGE_PREVIEW",
        },
    )
    assert blocked.status_code == 404
    assert "MESSAGE_GENERATION_NOT_READY" in blocked.json()["error"]["message"]
    knowledge_only = knowledge_client.post(
        "/api/v1/messages/generate",
        json={
            "format": "MT",
            "messageType": "MT998",
            "fields": [],
            "persist": False,
            "lane": "KNOWLEDGE_PREVIEW",
        },
    )
    assert knowledge_only.status_code == 404
    assert "STRUCTURE_SOURCE_MISSING" in knowledge_only.json()["error"]["message"]
    missing = knowledge_client.post(
        "/api/v1/messages/generate",
        json={
            "format": "MT",
            "messageType": "MT777",
            "fields": [],
            "persist": False,
            "lane": "KNOWLEDGE_PREVIEW",
        },
    )
    assert missing.status_code == 404
    assert "STRUCTURE_SOURCE_MISSING" in missing.json()["error"]["message"]


def test_the_preview_lane_is_never_implicit_and_the_configured_lane_is_unchanged(
    knowledge_client,
) -> None:  # type: ignore[no-untyped-def]
    # MT103 exists only as a preview pack; without the lane it is unknown, as before.
    assert (
        knowledge_client.get("/api/v1/messages/MT103/spec", params={"format": "MT"}).status_code
        == 404
    )
    assert (
        knowledge_client.post(
            "/api/v1/messages/generate",
            json={"format": "MT", "messageType": "MT103", "fields": [], "persist": False},
        ).status_code
        == 404
    )
    # MT541 configured resolves to the configured registry even though preview packs exist.
    spec = knowledge_client.get("/api/v1/messages/MT541/spec", params={"format": "MT"}).json()
    assert spec["lane"] == "CONFIGURED" and spec["release"] is None
    assert any(f["id"] == "SETTLEMENT-A-SEME" or f["qualifier"] == "SEME" for f in spec["fields"])
